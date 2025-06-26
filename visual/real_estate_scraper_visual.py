#import the needed libraries for the script
import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import csv
from webdriver_manager.chrome import ChromeDriverManager
import threading
import re
import time
import xlsxwriter
from datetime import datetime, timezone, timedelta
import imaplib
import email
from email.utils import parsedate_to_datetime
from dotenv import load_dotenv
import shutil
from openpyxl import load_workbook

def is_similar_grey(color_rgb):
    if not color_rgb or len(color_rgb) != 8:
        return False

    try:
        r = int(color_rgb[2:4], 16)
        g = int(color_rgb[4:6], 16)
        b = int(color_rgb[6:8], 16)
    except:
        return False

    # Base grey color (adjust if needed): e.g., RGB(191, 191, 191)
    base_r, base_g, base_b = 220, 220, 220
    tolerance = 15

    return (abs(r - base_r) <= tolerance and
            abs(g - base_g) <= tolerance and
            abs(b - base_b) <= tolerance)


def is_similar_red(color_rgb):
    if not color_rgb or len(color_rgb) != 8:
        return False

    try:
        r = int(color_rgb[2:4], 16)
        g = int(color_rgb[4:6], 16)
        b = int(color_rgb[6:8], 16)
    except:
        return False

    # Base red color from your Excel: RGB(217, 151, 149)
    base_r, base_g, base_b = 217, 151, 149
    tolerance = 15

    return (abs(r - base_r) <= tolerance and
            abs(g - base_g) <= tolerance and
            abs(b - base_b) <= tolerance)

def get_cleaned_numbers_string():
    # Load the Excel file
    df = pd.read_excel("data/input.xlsx")

    # Attempt to locate the column containing phone numbers
    number_col = next((col for col in df.columns if col.strip().lower() == 'phone number'), None)
    if not number_col:
        raise KeyError("Could not find a 'Phone Number' column in input.xlsx.")

    # Drop missing values and ensure all entries are strings
    numbers = df[number_col].dropna().astype(str)

    # Clean each number by removing non-digit characters
    cleaned_numbers = numbers.apply(lambda x: re.sub(r'\D', '', x))

    # Join all numbers with a space
    return ' '.join(cleaned_numbers)

def get_cleaned_numbers_list_only_red():
    # Load the Excel file
    df = pd.read_excel("data/reds.xlsx")

    # Attempt to locate the column containing phone numbers
    number_col = next((col for col in df.columns if col.strip().lower() == 'phone number'), None)
    if not number_col:
        raise KeyError("Could not find a 'Phone Number' column in reds.xlsx.")

    # Drop missing values and ensure all entries are strings
    numbers = df[number_col].dropna().astype(str)

    # Clean each number by removing non-digit characters
    cleaned_numbers = numbers.apply(lambda x: re.sub(r'\D', '', x))

    # Return the list of cleaned numbers
    return cleaned_numbers.tolist()

def get_cleaned_numbers_list():
    # Load the Excel file
    df = pd.read_excel("data/input.xlsx")

    # Attempt to locate the column containing phone numbers
    number_col = next((col for col in df.columns if col.strip().lower() == 'phone number'), None)
    if not number_col:
        raise KeyError("Could not find a 'Phone Number' column in input.xlsx.")

    # Drop missing values and ensure all entries are strings
    numbers = df[number_col].dropna().astype(str)

    # Clean each number by removing non-digit characters
    cleaned_numbers = numbers.apply(lambda x: re.sub(r'\D', '', x))

    # Return the list of cleaned numbers
    return cleaned_numbers.tolist()

def select_and_copy_file():
    filepath = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

    if not filepath:
        return

    try:
        # Load numbers from user-selected Excel file (headers on row 2)
        df_numbers = pd.read_excel(filepath, header=1)
        numbers = df_numbers[['Number']].dropna()
    except Exception as e:
        messagebox.showerror("Error", f"Could not read the selected file:\n{e}")
        return

    try:
        # Process numbers for input.xlsx using template
        template_path = "data/sample-phone-list.xlsx"
        df_template = pd.read_excel(template_path, header=None)

        df_template.iloc[1:, 0] = None
        df_template.iloc[1:, 1] = None

        for i, number in enumerate(numbers['Number'], start=1):
            df_template.at[i, 0] = number

        input_path = "data/input.xlsx"
        df_template.to_excel(input_path, index=False, header=False)

        # Process raw numbers for input2.csv with "Do Not Originate"
        cleaned_data = []
        for number in numbers['Number']:
            cleaned = re.sub(r'\D', '', str(number))  # Remove non-digit characters
            if cleaned:
                cleaned_data.append([f"{cleaned}", "Do Not Originate"])

        input2_path = "data/input2.csv"
        # Write to CSV file without header
        df_input2 = pd.DataFrame(cleaned_data, columns=["Phone Number", "Purpose"])
        df_input2.to_csv(input2_path, index=False, header=False)

        # Extract red-tagged numbers and write to reds.xlsx
        wb = load_workbook(filepath)
        ws = wb.active
        red_numbers = []

        for row in range(2, ws.max_row + 1):
            cell_b = ws.cell(row=row, column=2)
            cell_c = ws.cell(row=row, column=3)

            color_b = cell_b.fill.start_color.rgb if cell_b.fill and cell_b.fill.start_color else None
            color_c = cell_c.fill.start_color.rgb if cell_c.fill and cell_c.fill.start_color else None

            if is_similar_red(color_b) or is_similar_red(color_c) or is_similar_grey(color_b) or is_similar_grey(color_c):
                number_cell = ws.cell(row=row, column=1)
                number_value = number_cell.value
                if number_value:
                    red_numbers.append([number_value])

        if red_numbers:
            df_reds = pd.DataFrame(red_numbers, columns=["Phone Number"])
            df_reds.to_excel("data/reds.xlsx", index=False)

    except Exception as e:
        messagebox.showerror("Error", f"Processing failed:\n{e}")

def cleanup_inputs():
    for filename in ["data/input.xlsx", "data/input2.csv", "data/delete.csv", "data/reds.xlsx"]:
        try:
            if os.path.exists(filename):
                os.remove(filename)
            else:
                print(f"File not found: {filename}")
        except Exception as e:
            print(f"Error deleting {filename}: {e}")


def fetch_email_code_from(sender_email, start_time):
    timeout = 180  # 3 minutes
    interval = 10  # check every 10 seconds
    elapsed = 0

    while elapsed < timeout:
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        mail.select("inbox")

        result, data = mail.search(None, f'FROM "{sender_email}"')
        if result == "OK" and data[0]:
            mail_ids = data[0].split()
            latest_email_id = mail_ids[-1]
            result, msg_data = mail.fetch(latest_email_id, "(RFC822)")
            raw_email = msg_data[0][1]

            msg = email.message_from_bytes(raw_email)
            subject = msg["subject"]
            date_header = msg["date"]
            email_time = parsedate_to_datetime(date_header)

            if email_time >= start_time:
                match = re.search(r"\b\d{5}\b", subject)
                code = match.group()
                if match:
                    print(f"code is {code}")
                else:
                    print("Verification code not found in subject.")
                return code
            else:
                print(f"Email received before script started (received: {email_time}), waiting...")
        else:
            print("No message from sender yet, waiting...")

        time.sleep(interval)
        elapsed += interval

    print("No recent email received within the last 3 minutes.")
    return "none"

#this is the function that breaks up the excel file into the nuisance, nuisance_board_up and unsafe_structure CSVs
def process_file(file_path):
    #save the files in the data folder inside this directory
    os.makedirs("data", exist_ok=True)

    # Read the Excel file, skipping the first 6 rows that are the picture
    df = pd.read_excel(file_path, skiprows=6)

    # Normalize column names
    df.columns = df.columns.str.strip()

    # Define expected column names that we use in order to strip the desired pin numbers
    code_section_col = "Code Section"
    re_number_col = "RE Number"

    # Ensure required columns exist
    if code_section_col not in df.columns or re_number_col not in df.columns:
        messagebox.showerror("Error", "One or both expected columns ('Code Section', 'RE Number') are missing!")
        return None

    # Fill down missing values in "Code Section" because the code section only appears once per section
    #and not in each specific row
    df[code_section_col] = df[code_section_col].ffill()

    # Rename "RE Number" to "DocLegalDescription" in order for the other function to understand that
    # this is the section that refferes to PIN numbers for the properties
    df = df.rename(columns={re_number_col: "DocLegalDescription"})

    # Define categories and output filenames
    categories = {
        "Nuisance": "data/nuisance.csv",
        "Nuisance Board Up": "data/nuisance_board_up.csv",
        "Unsafe Structure": "data/unsafe_structure.csv",
        "Zoning-Commrc/Indstr": "data/commrc-indstr.csv"
    }

    output_files = {}
    for category, filename in categories.items():
        #get each pin for each category
        filtered_rows = df[df[code_section_col].str.strip().eq(category)]
        #drop any duplicates
        doc_legal_desc = filtered_rows[["DocLegalDescription"]].dropna().drop_duplicates()
        #write PIN + pin-number
        doc_legal_desc["DocLegalDescription"] = "PIN " + doc_legal_desc["DocLegalDescription"].astype(str)

        # Save to CSV with unique values
        doc_legal_desc.to_csv(filename, index=False)
        output_files[category] = filename

    return output_files

# Web scraping function for loading the wordlist.txt
def load_wordlist(filepath="data/wordlist.txt"):
    try:
        with open(filepath, "r", encoding="utf-8") as file:
            return [line.strip() for line in file if line.strip()]
    except FileNotFoundError:
        return []

#function using regular expression to check if a string is inside a word
#used mainly to check if anything in the wordlist is also in the scraped name, in order to
#detect if it is a business or person
def is_company(text, wordlist):
    pattern = r'\b(?:' + '|'.join(map(re.escape, wordlist)) + r')\b'
    return bool(re.search(pattern, text, re.IGNORECASE)) or "LLC" in text.upper()

#used to break up a scraped row into property address, zip and state
def parse_address_v1(address):
    match = re.match(r"(.+?)\s([A-Z]{2})\s(\d+)-?", address)
    if match:
        city, state, zip_code = match.groups()
        return city, state, zip_code
    return None

#used to break up a scraped row into mailing address, zip and state
def parse_address_v2(address):
    match = re.match(r"(.+),\s([A-Z]{2})\s(\d+)", address)
    if match:
        city, state, zip_code = match.groups()
        return city, state, zip_code
    return None

#used to break up row into first and last name
def get_first_last_name(text):
    words = text.split()
    last_name = words[0] if len(words) > 0 else ""
    first_name = words[1] if len(words) > 1 else ""
    return first_name, last_name

#gets the input csv and makes the list of urls depending on the pins
def make_urls(input_csv):
    with open(input_csv, mode='r', newline='', encoding='utf-8') as file:
        reader = list(csv.reader(file))

        if not reader:
            raise ValueError("CSV file is empty.")

        header = reader[0]

        if "DocLegalDescription" not in header:
            raise ValueError(f"Column 'DocLegalDescription' not found in the CSV file. Available columns: {header}")

        col_index = header.index("DocLegalDescription")

        # Extract all non-empty values in the column
        descriptions = [row[col_index] for row in reader[1:] if len(row) > col_index and row[col_index].strip()]

    pattern = re.compile(r'(\d{6})[-\s]+(\d{4})')

    extracted_urls = [
        f"https://paopropertysearch.coj.net/Basic/Detail.aspx?RE={match.group(1)}{match.group(2)}"
        for desc in descriptions if (match := pattern.search(desc))
    ]

    return extracted_urls

#the function to scrape a url from paopropertysearch
#get the wordlist, the url and the webdriver
def extract_data(url, wordlist, driver):
    #open the given url
    driver.get(url)
    #function to get text from an element if it exists, else return "Not found"
    def get_text(xpath):
        element = driver.find_elements("xpath", xpath)
        return element[0].text if element else "Not found"
    #get owner name
    owner_name = get_text('//span[contains(@id,"ctl00_cphBody_repeaterOwnerInformation_ctl00_lblOwnerName")]')
    #get mailing address
    mailing_address = get_text('//span[contains(@id,"ctl00_cphBody_repeaterOwnerInformation_ctl00_lblMailingAddressLine1")]')
    notes = ""
    #if c/o is in the mailing address then make the notes = mailing address
    #and set the mailing address as the second row
    if ("C/O" in mailing_address) or ("C/0" in mailing_address):
        notes = mailing_address
        mailing_address = get_text('//span[contains(@id,"ctl00_cphBody_repeaterOwnerInformation_ctl00_lblMailingAddressLine2")]')
    #if not then try to get the second line as notes, but keep them only if they are not empty
    else:
        notes = get_text('//span[contains(@id,"ctl00_cphBody_repeaterOwnerInformation_ctl00_lblMailingAddressLine2")]')
    if notes == "Not found":
        notes = ""
    #get mailing_city_info
    mailing_city_info = get_text('//span[contains(@id,"ctl00_cphBody_repeaterOwnerInformation_ctl00_lblMailingAddressLine3")]')
    #get property_address
    property_address = get_text('//span[contains(@id,"ctl00_cphBody_lblPrimarySiteAddressLine1")]')
    #get property_city_info
    property_city_info = get_text('//span[contains(@id,"ctl00_cphBody_lblPrimarySiteAddressLine2")]')
    #get re number
    re_number = get_text('//span[contains(@id,"ctl00_cphBody_lblRealEstateNumber")]')
    #get property_use
    property_use = get_text('//span[contains(@id,"ctl00_cphBody_lblPropertyUse")]')
    #try to find the table that contains the zoning info , if it doesn't exist return Not found
    try:
        table = driver.find_element(By.ID, "ctl00_cphBody_gridLand")
        rows = table.find_elements(By.TAG_NAME, "tr")
        second_row_columns = rows[1].find_elements(By.TAG_NAME, "td")
        zoning = second_row_columns[3].text
    except:
        zoning = "Not found"

    business_name, first_name, last_name = "", "", ""
    #split up mailing info into state, zip and address
    try:
        mailing_city, mailing_state, mailing_zip = parse_address_v2(mailing_city_info)
    except:
        mailing_city = mailing_state = mailing_zip = ""

    #split up property info into state, zip and address
    try:
        property_city, property_state, property_zip = parse_address_v1(property_city_info)
    except:
        property_city = property_state = property_zip = ""
    #get first and last name if it is not a company
    try:
        business_name = owner_name
        if (not is_company(owner_name, wordlist)):
            first_name, last_name = get_first_last_name(owner_name)
    except Exception as e:
        print("Error:", e)

    return [business_name, first_name, last_name, mailing_address, mailing_city, mailing_state, mailing_zip,
            property_address, property_city, property_state, property_zip, url, re_number, property_use, zoning, notes]

#make the headless webdriver and scrape the urls
def start_scraping(input_csv, stop_flag, progress_var, progress_label, status_label, total_urls, progress_step):
    #webdriver options
    options = Options()
    options.add_argument('--headless')
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    #get the urls and write the first row of the excel output
    url_list = make_urls(input_csv)
    file_path = 'business_info.xlsx'
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet('Business Info')

    headers = [
        "Business_name", "first_name", "last_name", "mailing_address", "mailing_city", "mailing_state", "mailing_zip",
        "property_address", "property_city", "property_state", "property_zip", "link", "RE #", "Property use", "Zoning Assesment", "Notes"
    ]
    worksheet.write_row(0, 0, headers)

    wordlist = load_wordlist()
    x_all = len(url_list)
    #scrape the urls and update the progress bar
    try:
        for index, url in enumerate(url_list, start=1):
            if stop_flag[0]:
                print("Scraping stopped by user.")
                status_label.config(text="Stopped", fg="red")
                break
            data = extract_data(url, wordlist, driver)
            worksheet.write_row(index, 0, data)
            print(f"{url} - {index} out of {x_all}")

            progress_var.set(index * progress_step)
            progress_label.config(text=f"{index} out of {x_all}")
            status_label.config(text="Running", fg="green")
            root.update_idletasks()
    except Exception as e:
        print(f"An error occurred: {e}")
    #save the excel file
    driver.quit()
    workbook.close()
    print("Scraping completed and Excel file created!")
    status_label.config(text="Stopped", fg="red")

#upload file for the nuisance, nuisance_board_up and unsafe_structure CSVs
def upload_file(status_label):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return

    status_label.config(text="Processing file...")
    output_files = process_file(file_path)

    if output_files:
        status_label.config(text="Processing completed! Download files from 'data' folder.")
        messagebox.showinfo("Success", "CSV files have been generated successfully!")
    else:
        status_label.config(text="Processing failed.")

#select file for the scraping
def select_file_for_scraping():
    file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    return file_path

#function to download the LIEN csv
def download_LIEN_data():
    #save in the data folder in this directory
    download_folder = os.path.join(os.getcwd(), "data")
    #webdriver options
    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_experimental_option("detach", True)
    options.add_argument('--headless')
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-features=VizDisplayCompositor")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    # Set download preferences for Chrome
    prefs = {
        "download.default_directory": download_folder,  # Set the download directory
        "download.prompt_for_download": False,  # Disable the download prompt
        "directory_upgrade": True
    }

    options.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    #open duval clerk
    driver.get("https://or.duvalclerk.com/Search/Disclaimer?st=/search/SearchTypeDocType")
    #accept the conditions
    submit_button = driver.find_element(By.ID, "btnButton")
    submit_button.click()
    #click DocTypesDisplay
    input_field = driver.find_element(By.ID, "DocTypesDisplay-input")
    input_field.clear()  # Clear the field in case there is any default value

    # Write "LIEN (LN)" into the input field
    input_field.send_keys("LIEN (LN)")

    # Find the span element by its text and click it
    span_element = driver.find_element(By.XPATH, "//span[contains(text(),'Specify Date Range...')]")
    span_element.click()

    # Wait for the dropdown to appear
    time.sleep(1)  # Adjust timing if necessary

    # Find the li element by its text and click the Last 90 Days option
    li_element = driver.find_element(By.XPATH, "//li[contains(text(),'Last 90 Days')]")
    li_element.click()
    time.sleep(1)
    #submit these parameters
    submit_button2 = driver.find_element(By.ID, "btnSearch")
    submit_button2.click()
    time.sleep(3.5)
    #find the input field
    input_element = driver.find_element(By.ID, "fldText")
    input_element.clear()  # Clear the field if there is any pre-filled text
    # Write "Jacksonville city of" into the input field
    input_element.send_keys("Jacksonville city of")
    #Find search button and click it
    button_elements = driver.find_elements(By.CLASS_NAME, "t-button")
    button_elements[4].click()
    time.sleep(1)
    #click download button
    button_elements[12].click()
    print(f"File should have been downloaded to: {download_folder}")

#function to download the tax deed sale csv
def download_tax_deed_sale_data():
    #save in the data folder in this directory
    download_folder = os.path.join(os.getcwd(), "data")
    #webdriver options
    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_experimental_option("detach", True)
    options.add_argument('--headless')
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-features=VizDisplayCompositor")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    # Set download preferences for Chrome
    prefs = {
        "download.default_directory": download_folder,  # Set the download directory
        "download.prompt_for_download": False,  # Disable the download prompt
        "directory_upgrade": True
    }

    options.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    #open duval clerk
    driver.get("https://or.duvalclerk.com/Search/Disclaimer?st=/search/SearchTypeDocType")
    #accept the conditions
    submit_button = driver.find_element(By.ID, "btnButton")
    submit_button.click()
    #click DocTypesDisplay
    input_field = driver.find_element(By.ID, "DocTypesDisplay-input")
    input_field.clear()  # Clear the field in case there is any default value

    # Write "LIEN (LN)" into the input field
    input_field.send_keys("NOTICE OF TAX DEED SALE (NTD)")

    # Find the span element by its text and click it
    span_element = driver.find_element(By.XPATH, "//span[contains(text(),'Specify Date Range...')]")
    span_element.click()

    # Wait for the dropdown to appear
    time.sleep(1)  # Adjust timing if necessary

    # Find the li element by its text and click the Last 90 Days option
    li_element = driver.find_element(By.XPATH, "//li[contains(text(),'Last 30 Days')]")
    li_element.click()
    time.sleep(1)
    #submit these parameters
    submit_button2 = driver.find_element(By.ID, "btnSearch")
    submit_button2.click()
    time.sleep(3.5)
    #Find search button and click it
    button_elements = driver.find_elements(By.CLASS_NAME, "t-button")
    #click download button
    button_elements[12].click()
    print(f"File should have been downloaded to: {download_folder}")

def apply_to_remove_scam_likely_status():
    select_and_copy_file()  # always runs

    if run_free_caller.get():
        free_caller_registry()
    if run_tns.get():
        tns()
    if run_hiya.get():
        hiya()
    if run_first_orion.get():
        first_orion()
    if run_tmobile.get():
        tmobile()
    if run_verizon.get():
        verizon()

    cleanup_inputs()  # always runs


def verizon():
    print("doing verizon")
    options = Options()
    # options.add_argument('--headless')
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    driver.get("https://voicespamfeedback.com/vsf/")
    wait = WebDriverWait(driver, 30)  # Adjust timeout as needed

    # Click the first radio button
    first_radio = wait.until(EC.element_to_be_clickable((By.ID, "type_enterprise")))
    first_radio.click()

    # Click the second radio button
    second_radio = wait.until(EC.element_to_be_clickable((By.ID, "type-enterprise")))
    second_radio.click()

    # Wait until the "Next" button is clickable
    next_button = wait.until(EC.element_to_be_clickable((By.ID, "nextButton")))

    # Click the "Next" button
    next_button.click()

    # Enter phone number
    nums = get_cleaned_numbers_list()
    phone_input = driver.find_element(By.ID, "enterprise_phone_0")
    phone_input.clear()
    phone_input.send_keys(nums[0])

    # Upload file
    file_input = driver.find_element(By.ID, "enterpriseFile")
    file_path = os.path.abspath("./data/input.xlsx")
    file_input.send_keys(file_path)

    # Select 'Other' from dropdown
    category_dropdown = Select(driver.find_element(By.ID, "enterprise_category"))
    category_dropdown.select_by_value("other")

    # Fill contact name
    contact_name = driver.find_element(By.ID, "enterprise_contact_name")
    contact_name.send_keys("Denny Anderson")

    # Fill contact phone
    contact_phone = driver.find_element(By.ID, "enterprise_contact_phone")
    contact_phone.send_keys("3523596700")

    # Fill contact email
    contact_email = driver.find_element(By.ID, "enterprise_contact_email")
    contact_email.send_keys("dennyanderson72@gmail.com")

    # Fill company name
    company_name = driver.find_element(By.ID, "enterprise_company_name")
    company_name.send_keys("Built to Buy LLC")

    # Company Address Line 1
    address1 = driver.find_element(By.ID, "enterprise_company_address")
    address1.send_keys("1031 1st st s #806, 806")

    # Service Provider
    service_provider = driver.find_element(By.ID, "enterprise_service_provider")
    service_provider.send_keys("Onvoy, Twilio")

    # Call Count
    call_count = driver.find_element(By.ID, "call_count")
    call_count.send_keys("3000")
    # time.sleep(60)
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((
            By.XPATH,
            "//ul[@role='tablist']/li[3][contains(@class, 'current') and @aria-selected='true']"
        ))
    )
    time.sleep(2)

def tmobile():
    print("doing t-mobile")
    options = Options()
    # options.add_argument('--headless')
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    wait = WebDriverWait(driver, 10)
    wait2 = WebDriverWait(driver, 300)
    if only_red.get():
        nums = get_cleaned_numbers_list_only_red()
    else:
        nums = get_cleaned_numbers_list()
    for num in nums:
        driver.get("https://callreporting.t-mobile.com/")
        elements = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "v-input--selection-controls__ripple")))
        elements[0].click()
        elements = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "v-input--selection-controls__ripple")))
        elements[2].click()
        # Input into the first field (input-18)
        input1 = driver.find_element(By.ID, "input-18")

        # Scroll into view
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", input1)
        time.sleep(0.5)  # Allow scroll animation to complete

        # Input text
        input1.send_keys(num)  # Replace with your value

        # Input into the second field (input-21)
        input2 = driver.find_element(By.ID, "input-21")
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", input2)
        time.sleep(0.5)

        input2.send_keys(num)  # Replace with your value
        # Find the span with text "Review"
        review_span = wait.until(
            EC.presence_of_element_located((By.XPATH, "//span[text()='Review']"))
        )
        time.sleep(2)
        # Click the parent button (recommended)
        review_button = review_span.find_element(By.XPATH, "./ancestor::button")
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", review_button)
        wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='Review']/ancestor::button"))).click()
        # XPath to target the 'Submit report' button by visible text
        submit_xpath = "//span[text()='Submit report']/ancestor::button"

        # Wait until the button is enabled and clickable
        wait2.until(lambda d: (
            (el := d.find_element(By.XPATH, submit_xpath)) and
            el.is_enabled() and
            "v-btn--disabled" not in el.get_attribute("class")
        ))

        # Click the button
        driver.find_element(By.XPATH, submit_xpath).click()
        time.sleep(2)

def free_caller_registry():
    print("doing free-caller-Registry")
    start_time = datetime.now(timezone.utc).replace(second=0, microsecond=0)
    #webdriver options
    options = Options()
    #options.add_argument('--headless')
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    driver.get("https://freecallerregistry.com/fcr/")
    try:
        # Wait for the button to be clickable and click it
        time.sleep(3)  # Consider replacing with WebDriverWait for more robustness
        button = driver.find_element(By.ID, "nextButton")
        button.click()
    except Exception as e:
        print(f"Error clicking button: {e}")

    # Fill contact email
    contact_email = driver.find_element(By.ID, "enterprise_contact_email")
    contact_email.send_keys("dennyanderson72@gmail.com")


    # Click the "Send Verification Code" button
    send_code_btn = driver.find_element(By.ID, "send-verification-code")
    send_code_btn.click()

    # Enter phone number
    nums = get_cleaned_numbers_list()
    phone_input = driver.find_element(By.ID, "enterprise_phone_0")
    phone_input.clear()
    phone_input.send_keys(nums[0])

    # Upload file
    file_input = driver.find_element(By.ID, "enterpriseFile")
    file_path = os.path.abspath("./data/input.xlsx")
    file_input.send_keys(file_path)

    # Select 'Other' from dropdown
    category_dropdown = Select(driver.find_element(By.ID, "enterprise_category"))
    category_dropdown.select_by_value("other")

    # Fill contact name
    contact_name = driver.find_element(By.ID, "enterprise_contact_name")
    contact_name.send_keys("Denny Anderson")

    # Fill contact phone
    contact_phone = driver.find_element(By.ID, "enterprise_contact_phone")
    contact_phone.send_keys("3523596700")

    # Fill company name
    company_name = driver.find_element(By.ID, "enterprise_company_name")
    company_name.send_keys("Built to Buy LLC")

    # Company Address Line 1
    address1 = driver.find_element(By.ID, "enterprise_company_address_line_1")
    address1.send_keys("1031 1st st s #806")

    # Company Address Line 2
    address2 = driver.find_element(By.ID, "enterprise_company_address_line_2")
    address2.send_keys("806")

    # City
    city = driver.find_element(By.ID, "enterprise_company_address_city")
    city.send_keys("Jacksonville Beach")

    # State (Select)
    state_select = Select(driver.find_element(By.ID, "enterprise_company_address_state"))
    state_select.select_by_value("FL")

    # ZIP Code
    zip_code = driver.find_element(By.ID, "enterprise_company_address_zip")
    zip_code.send_keys("32250")

    # Company URL
    company_url = driver.find_element(By.ID, "enterprise_company_url")
    company_url.send_keys("https://builttobuy.org/")

    # Service Provider
    service_provider = driver.find_element(By.ID, "enterprise_service_provider")
    service_provider.send_keys("Onvoy, Twilio")

    # Call Count
    call_count = driver.find_element(By.ID, "call_count")
    call_count.send_keys("3000")

    #verify email code
    code = fetch_email_code_from("no-reply@tnsi.com",start_time)
    captcha_input = driver.find_element(By.ID, "captcha")
    captcha_input.clear()
    captcha_input.send_keys(code)

    # Click the "Submit" button
    submit_btn = driver.find_element(By.ID, "submitButton")
    submit_btn.click()


def tns():
    print("doing tns")
    start_time = datetime.now(timezone.utc).replace(second=0, microsecond=0)
    #webdriver options
    options = Options()
    #options.add_argument('--headless')
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    driver.get("https://reportarobocall.com/trf/#")

    radio_btn = driver.find_element(By.ID, "type_enterprise")
    radio_btn.click()

    time.sleep(1)
    radio_btn = driver.find_element(By.ID, "claim_enterprise")
    driver.execute_script("arguments[0].click();", radio_btn)


    time.sleep(1)
    next_button = driver.find_element(By.ID, "nextButton")
    next_button.click()

    # Fill contact email
    contact_email = driver.find_element(By.ID, "enterprise_contact_email")
    contact_email.send_keys("dennyanderson72@gmail.com")

    # Click the "Send Verification Code" button
    send_code_btn = driver.find_element(By.ID, "send-verification-code")
    send_code_btn.click()

    # Enter phone number
    nums = get_cleaned_numbers_list()
    phone_input = driver.find_element(By.ID, "enterprise_phone_0")
    phone_input.clear()
    phone_input.send_keys(nums[0])

    # Upload file
    file_input = driver.find_element(By.ID, "enterpriseFile")
    file_path = os.path.abspath("./data/input.xlsx")
    file_input.send_keys(file_path)

    # Select 'Other' from dropdown
    category_dropdown = Select(driver.find_element(By.ID, "enterprise_category"))
    category_dropdown.select_by_value("other")

    # Fill contact name
    contact_name = driver.find_element(By.ID, "enterprise_contact_name")
    contact_name.send_keys("Denny Anderson")

    # Fill contact phone
    contact_phone = driver.find_element(By.ID, "enterprise_contact_phone")
    contact_phone.send_keys("3523596700")

    # Fill company name
    company_name = driver.find_element(By.ID, "enterprise_company_name")
    company_name.send_keys("Built to Buy LLC")

    # Company Address Line 1
    address1 = driver.find_element(By.ID, "enterprise_company_address")
    address1.send_keys("1031 1st st s #806, 806")

    # Service Provider
    service_provider = driver.find_element(By.ID, "enterprise_service_provider")
    service_provider.send_keys("Onvoy, Twilio")

    # Call Count
    call_count = driver.find_element(By.ID, "call_count")
    call_count.send_keys("3000")

    #verify email code
    code = fetch_email_code_from("no-reply@tnsi.com",start_time)
    captcha_input = driver.find_element(By.ID, "captcha")
    captcha_input.clear()
    captcha_input.send_keys(code)

    # Click the "Submit" button
    submit_btn = driver.find_element(By.ID, "submitButton")
    submit_btn.click()


def hiya():
    print("doing hiya")
    #webdriver options
    options = Options()
    #options.add_argument('--headless')
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    driver.get("https://hiyahelp.zendesk.com/hc/en-us/p/screening/")

    h3_element = driver.find_element(By.XPATH, "//h3[text()='I’m a business.']")

    # Go to its parent, then grandparent
    parent = h3_element.find_element(By.XPATH, "..")

    # Find the only button child within 'screen-module_copy' and click it
    button = parent.find_element(By.XPATH, ".//button")
    button.click()

    h3_element = driver.find_element(By.XPATH, "//h3[text()='My own number(s)']")

    # Go to its parent, then grandparent
    parent = h3_element.find_element(By.XPATH, "..")

    # Find the only button child within 'screen-module_copy' and click it
    button = parent.find_element(By.XPATH, ".//button")
    button.click()

    # Locate the div element by its class
    screening_div = driver.find_element(By.CLASS_NAME, 'screening_module.yes-account')

    # Within that div, locate the 'a' element and click it
    report_problem_link = screening_div.find_element(By.TAG_NAME, 'a')
    report_problem_link.click()

    time.sleep(3)
    # Locate the email input field and send keys to it
    email_input = driver.find_element(By.ID, 'request_anonymous_requester_email')
    email_input.send_keys("dennyanderson72@gmail.com")

    # Locate the description textarea and send keys to it
    description_textarea = driver.find_element(By.ID, 'request_description')
    description_textarea.send_keys("My numbers for Built to Buy LLC are being mislabeled as spam. All our calls are opted in and we follow all tns, hiya and guardian regulation guidelines")

    # Locate the "Business" option (hidden input and the <a> element) and click on it
    option = driver.find_element(By.ID, "request_custom_fields_81018947")
    driver.execute_script("arguments[0].value = 'business';", option)
    time.sleep(1)

    # Find the first input element by its ID and send text
    input1 = driver.find_element(By.ID, "request_custom_fields_81019067")
    input1.send_keys("Built to Buy LLC")

    # Find the second input element by its ID and send text
    input2 = driver.find_element(By.ID, "request_custom_fields_81020527")
    input2.send_keys("https://builttobuy.org/")

    # Find the third input element by its ID and send text
    input3 = driver.find_element(By.ID, "request_custom_fields_81018847")
    data = get_cleaned_numbers_string()
    input3.send_keys(data)

    # Find the fourth input element by its ID and send text
    input4 = driver.find_element(By.ID, "request_custom_fields_32855858154387")
    input4.send_keys("Denny Anderson")

    # Wait until the checkbox is clickable
    checkbox = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "request_custom_fields_32902380019859"))
    )

    # Scroll into view and click
    driver.execute_script("arguments[0].scrollIntoView(true);", checkbox)
    checkbox.click()

    # Wait for the submit button in the footer to be clickable
    submit_button_footer = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//footer//input[@type='submit' and @name='commit']"))
    )

    # Scroll to the footer if needed
    driver.execute_script("arguments[0].scrollIntoView(true);", submit_button_footer)

    # Click the submit button in the footer
    submit_button_footer.click()

def first_orion():
    print("doing first-orion")
    options = Options()
    #options.add_argument('--headless')
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    driver.get("https://portal.firstorion.com/app/landing#/login")
    # Locate input fields and send keys
    driver.find_element(By.ID, "email-input").send_keys(ORION_ADDRESS)
    driver.find_element(By.ID, "password-input").send_keys(ORION_PASSWORD)

    # Click the login button
    driver.find_element(By.CSS_SELECTOR, "button.btn.btn-md.btn-primary.btn-block").click()
    wait = WebDriverWait(driver, 30)
    phone_numbers_button = wait.until(
        EC.element_to_be_clickable((By.ID, "phone-numbers-nav-item"))
    )
    phone_numbers_button.click()
    time.sleep(10)
    # Loop until the button no longer exists
    while True:
        try:
            # Wait for the parent div to appear
            parent_div = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.row.align-items-center.custom-pagination"))
            )

            # Now find the child button inside the parent div
            load_more_button = parent_div.find_element(By.CSS_SELECTOR, "button.btn.btn-primary.font-weight-bold.btn-sm.d-flex.align-items-center.justify-content-center")

            # Scroll to the button to ensure it's in view
            driver.execute_script("arguments[0].scrollIntoView(true);", load_more_button)
            time.sleep(1)  # Optional: Wait for the scroll action to complete

            # Click the "Load More" button
            load_more_button.click()
        except:
            break
    # Find all <tr> rows with phone number data
    rows = driver.find_elements(By.XPATH, "//tr[contains(@class, 'hover:tw-cursor-pointer')]")

    # Data storage
    data = []

    # Extract phone numbers and purposes
    for row in rows:
        try:
            number = row.find_element(By.CLASS_NAME, "phonenumber").text.strip()
            purpose = row.find_element(By.CLASS_NAME, "phonenumber-call-purpose").text.strip()
            data.append({"Phone Number": number, "Purpose": purpose})
        except Exception as e:
            print("Skipping row due to missing data:", e)

    # Create DataFrame
    df = pd.DataFrame(data)

    # Save to CSV file
    output_path = os.path.join("data", "delete.csv")
    df.to_csv(output_path, index=False)

     # Step 1: Click on the "create-dropdown" button
    create_dropdown_button = WebDriverWait(driver, 15).until(
        EC.element_to_be_clickable((By.ID, "create-dropdown"))
    )
    # Scroll to the button to ensure it's in view
    driver.execute_script("arguments[0].scrollIntoView(true);", create_dropdown_button)
    time.sleep(1)  # Optional: Wait for the scroll action to complete
    create_dropdown_button.click()

    dropdown_menu = driver.find_element(By.CSS_SELECTOR, "ul.dropdown-menu.show")
    # Find the second <li> element (Delete Phone Numbers) and click the <a> tag inside it
    second_li = dropdown_menu.find_elements(By.TAG_NAME, "li")[1]  # Index 1 for the second <li>
    delete_phone_numbers_link = second_li.find_element(By.TAG_NAME, "a")

    # Click the "Delete Phone Numbers" link
    delete_phone_numbers_link.click()

    try:
        # Wait for the nav tab list to be visible
        nav_tabs = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "ul.nav.nav-tabs[role='tablist']"))
        )

        # Get all <li> elements inside the <ul>
        tab_items = nav_tabs.find_elements(By.TAG_NAME, "li")

        # Access the second <li> (index 1) and find the <button> inside it
        upload_csv_button = tab_items[1].find_element(By.TAG_NAME, "button")

        # Click the button
        upload_csv_button.click()
    except Exception as e:
        print(f"Failed to click the second tab: {e}")

    file_path = os.path.abspath("data/delete.csv")
    try:
        # Wait until the file input is present (likely inside the same container as the button)
        file_input = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']"))
        )

        # Upload the file by sending the file path
        file_input.send_keys(file_path)
        print("File uploaded successfully.")
    except Exception as e:
        print(f"Failed to upload file: {e}")

    try:
        # Wait until the Delete button is clickable
        delete_button = WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'btn-outline-dark') and text()='Delete']"))
        )
        delete_button.click()
    except Exception as e:
        print(f"Failed to click 'Delete' button: {e}")
    time.sleep(4)
    # Wait for the "Close" button to be clickable
    close_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "button.btn-close[aria-label='Close']"))
    )

    # Click the "Close" button
    close_button.click()

    # Wait for the "Add Phone Numbers" button to be clickable
    add_phone_numbers_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//button[text()='Add Phone Numbers']"))
    )

    # Click the "Add Phone Numbers" button
    add_phone_numbers_button.click()

    try:
        # Wait for the nav tab list to be visible
        nav_tabs = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "ul.nav.nav-tabs[role='tablist']"))
        )

        # Get all <li> elements inside the <ul>
        tab_items = nav_tabs.find_elements(By.TAG_NAME, "li")

        # Access the second <li> (index 1) and find the <button> inside it
        upload_csv_button = tab_items[1].find_element(By.TAG_NAME, "button")

        # Click the button
        upload_csv_button.click()
    except Exception as e:
        print(f"Failed to click the second tab: {e}")

    file_path = os.path.abspath("data/input2.csv")
    try:
        # Wait until the file input is present (likely inside the same container as the button)
        file_input = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']"))
        )

        # Upload the file by sending the file path
        file_input.send_keys(file_path)
        print("File uploaded successfully.")
    except Exception as e:
        print(f"Failed to upload file: {e}")

    # Wait for the second "Save" button to be clickable
    save_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//div[@class='col text-right']//button[2]"))
    )

    # Click the "Save" button
    save_button.click()

    time.sleep(4)
    # Wait for the "Close" button to be clickable
    close_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "button.btn-close[aria-label='Close']"))
    )

    # Click the "Close" button
    close_button.click()

def download_lien_data():
    status_label.config(text="Started downloading LIEN data...")
    threading.Thread(target=download_LIEN_data).start()

def download_tax_deed_data():
    status_label.config(text="Started downloading Tax Deed Sale data...")
    threading.Thread(target=download_tax_deed_sale_data).start()

def start_selected_download():
    selection = download_var.get()
    if selection == "LIEN":
        download_lien_data()
    elif selection == "Tax Deed Sale":
        download_tax_deed_data()
    else:
        status_label.config(text="Please select a download option.")

def run_gui():
    #this flag is to stop the scraping if you press the stop button
    stop_flag = [False]
    #pressinf stop make the flag true
    def on_stop_button_click():
        stop_flag[0] = True

    def on_start_button_click():
        stop_flag[0] = False
        input_csv = select_file_for_scraping()

        if input_csv:
            total_urls = len(make_urls(input_csv))
            progress_step = 100 / total_urls if total_urls > 0 else 100

            threading.Thread(target=start_scraping, args=(input_csv, stop_flag, progress_var, progress_label, status_label, total_urls, progress_step)).start()


    global root, status_label, download_var
    root = tk.Tk()

    global only_red, run_free_caller, run_tns, run_hiya, run_first_orion, run_tmobile, run_verizon
    # Checkbox control variables
    run_free_caller = tk.BooleanVar(value=True)
    run_tns = tk.BooleanVar(value=True)
    run_hiya = tk.BooleanVar(value=True)
    only_red = tk.BooleanVar(value=True)
    run_first_orion = tk.BooleanVar(value=True)
    run_tmobile = tk.BooleanVar(value=False)
    run_verizon = tk.BooleanVar(value=True)

    root.title("Housing Scraper")

    title_label = tk.Label(root, text="        Real Estate Scraper and Appraiser       ", font=("Arial", 20))
    title_label.pack(pady=10)

    upload_button = tk.Button(root, text="Jax code split up", command=lambda: upload_file(status_label), font=("Arial", 12))
    upload_button.pack(pady=5)

    scraping_button = tk.Button(root, text="Start Scraping", command=on_start_button_click, font=("Arial", 12))
    scraping_button.pack(pady=5)

    stop_button = tk.Button(root, text="Stop Scraping", command=on_stop_button_click, font=("Arial", 12))
    stop_button.pack(pady=5)

    # Dropdown for selecting download type
    download_var = tk.StringVar()
    download_label = tk.Label(root, text="Download Data:", font=("Arial", 12))
    download_label.pack(pady=5)

    download_dropdown = ttk.Combobox(root, textvariable=download_var, state="readonly", font=("Arial", 12))
    download_dropdown['values'] = ("LIEN", "Tax Deed Sale")
    download_dropdown.pack(pady=5)

    # Button to start download
    download_button = tk.Button(root, text="Start Download", command=start_selected_download, font=("Arial", 12))
    download_button.pack(pady=5)

    apply_button = tk.Button(root, text="Apply to Remove 'Scam Likely' Status", command=apply_to_remove_scam_likely_status, font=("Arial", 12))
    apply_button.pack(pady=5)

    # Checkboxes to enable/disable each site
    tk.Checkbutton(root, text="Free Caller Registry", variable=run_free_caller).pack(anchor='w', padx=20)
    tk.Checkbutton(root, text="TNS", variable=run_tns).pack(anchor='w', padx=20)
    tk.Checkbutton(root, text="Hiya", variable=run_hiya).pack(anchor='w', padx=20)
    tk.Checkbutton(root, text="First Orion", variable=run_first_orion).pack(anchor='w', padx=20)
    tmobile_frame = tk.Frame(root)
    tmobile_frame.pack(anchor='w', padx=20)
    tk.Checkbutton(tmobile_frame, text="T-Mobile", variable=run_tmobile).pack(side='left')
    tk.Checkbutton(tmobile_frame, text="only use flagged numbers", variable=only_red).pack(side='left', padx=10)
    tk.Checkbutton(root, text="Verizon", variable=run_verizon).pack(anchor='w', padx=20)

    status_label = tk.Label(root, text="", fg="blue", font=("Arial", 10))
    status_label.pack(pady=10)

    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(root, variable=progress_var, maximum=100)
    progress_bar.pack(padx=10, pady=10, fill=tk.X)

    progress_label = tk.Label(root, text="0 out of 0")
    progress_label.pack(padx=10, pady=10)

    exit_button = tk.Button(root, text="Exit", command=root.quit, font=("Arial", 12))
    exit_button.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    # Load environment variables
    load_dotenv()

    EMAIL_ADDRESS = os.getenv("EMAIL_ADDRESS")
    EMAIL_PASSWORD = os.getenv("EMAIL_PASSWORD")
    ORION_ADDRESS = os.getenv("ORION_ADDRESS")
    ORION_PASSWORD = os.getenv("ORION_PASSWORD")
    run_gui()

